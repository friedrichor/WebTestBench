import argparse
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook

from utils import *


class ScoringPipeline_Oracle:
    """
    End-to-end scorer for WebProberBench outputs (gold-alignment variant).

    Responsibilities:
    1. Load dataset records and corresponding agent outputs.
    2. Parse gold checklist items and predicted checklist items.
    3. Align predictions to gold items directly by checklist order.
    4. Compute precision, recall, and F1 metrics.
    5. Aggregate metrics and export JSON/Excel reports.
    """

    def __init__(
        self,
        dataset_path: Path,
        output_root: Path,
        version: str,
    ) -> None:
        self.dataset_path = dataset_path
        self.output_root = output_root
        self.version = version
        self.dataset = self._load_dataset()

        # Task IDs with missing extracted result files.
        self.missing_result_ids: List[str] = []
        # Task IDs where predicted and gold checklist lengths differ.
        self.mismatch_item_ids: List[str] = []
        # Task IDs where all predicted checklist items are marked failed.
        self.all_fail_pred_ids: List[str] = []

        # Stable ordering used when rendering grouped reports.
        self.ordered_categories = [
            "Presentation", "Search", "Tool", "Commerce",
            "Data Management", "Workflow", "User-Generated Content",
        ]
        self.ordered_classes = ["functionality", "constraint", "interaction", "content"]

    def run(self) -> None:
        """
        Execute the full scoring workflow:
        1. Iterate through all dataset records.
        2. Score each record independently.
        3. Aggregate metrics and write summary artifacts.
        """
        # Clear stale missing-result logs from prior runs.
        missing_path = self.output_root / "missing_results.json"
        if missing_path.exists():
            missing_path.unlink()

        # Initialize in-memory accumulators.
        aggregators = self._initialize_aggregators()
        
        # Score each record and fold metrics into accumulators.
        for record_id, record in self.dataset.items():
            aggregators['total_count'] += 1
            
            output_dir = self.output_root / record_id
            if not output_dir.exists():
                print(f"Warning: output dir not found for {record_id}, skipping.")
                continue
                
            print_green(f"Scoring {record_id}")
            
            # Count gold checklist items per class.
            self._update_class_item_counts(record, aggregators['class_item_counts'])
            
            # Process one record and return its scoring bundle.
            result_bundle = self._process_record(record_id, record, output_dir)
            if result_bundle is None:
                continue
                
            # Merge per-record metrics into global totals.
            self._update_aggregators(result_bundle, record, aggregators)

        # Persist missing-result task IDs for auditing.
        if self.missing_result_ids:
            self._write_json(missing_path, {"missing_result": self.missing_result_ids})

        # Compute final averages and write artifacts.
        self._compute_and_save_final_results(aggregators)
        
        # Report checklist-size mismatches.
        if self.mismatch_item_ids:
            print_red(f"Checklist count mismatch for: {self.mismatch_item_ids}")
        
        # Report tasks where all predicted items failed.
        if self.all_fail_pred_ids:
            print_red(f"All pred items failed for: {self.all_fail_pred_ids}")

    def _initialize_aggregators(self) -> Dict:
        """Create and return all aggregate counters used across records."""
        return {
            'total_count': 0,
            'scored_count': 0,
            'scored_count_no_missing': 0,
            'total_metrics': {"precision": 0.0, "recall": 0.0, "f1": 0.0},
            'total_metrics_no_missing': {"precision": 0.0, "recall": 0.0, "f1": 0.0},
            'category_totals': {},
            'category_counts': {},
            'class_totals': {},
            'class_counts': {},
            'class_item_counts': {cls: 0 for cls in self.ordered_classes},
        }

    def _update_class_item_counts(self, record: dict, class_item_counts: Dict[str, int]) -> None:
        """Count gold checklist items per class for denominator reporting."""
        gold_items = self._parse_gold_checklist(record)
        for item in gold_items.values():
            cls = item.get("class")
            if cls in class_item_counts:
                class_item_counts[cls] += 1

    def _update_aggregators(self, result_bundle: dict, record: dict, aggregators: Dict) -> None:
        """Update all global aggregates using one processed record bundle."""
        metrics = result_bundle.get("overall")
        class_metrics = result_bundle.get("by_class", {})
        match_ids = result_bundle.get("match_ids")
        gold_items = result_bundle.get("gold_items", {})
        missing_result = metrics.get("missing_result", False)

        # Update overall metrics (includes missing cases).
        self._accumulate_metrics(aggregators['total_metrics'], metrics)
        aggregators['scored_count'] += 1
        
        # Update "no_missing" slice.
        if not missing_result:
            self._accumulate_metrics(aggregators['total_metrics_no_missing'], metrics)
            aggregators['scored_count_no_missing'] += 1

        # Update category-level aggregates.
        self._update_category_stats(record, metrics, aggregators)
        
        # Update class-level aggregates.
        self._update_class_stats(class_metrics, aggregators)
        
    def _accumulate_metrics(self, target: Dict[str, float], source: Dict[str, float]) -> None:
        """Add values from `source` metrics into `target` metrics in place."""
        for key in target:
            target[key] += float(source.get(key, 0.0))

    def _update_category_stats(self, record: dict, metrics: Dict, aggregators: Dict) -> None:
        """Accumulate metrics grouped by task category."""
        category = record.get("category", "Unknown")
        
        if category not in aggregators['category_totals']:
            aggregators['category_totals'][category] = {k: 0.0 for k in aggregators['total_metrics']}
            aggregators['category_counts'][category] = 0
        
        self._accumulate_metrics(aggregators['category_totals'][category], metrics)
        aggregators['category_counts'][category] += 1

    def _update_class_stats(self, class_metrics: Dict, aggregators: Dict) -> None:
        """Accumulate metrics grouped by checklist class."""
        for cls in self.ordered_classes:
            cls_metrics = class_metrics.get(cls)
            if not cls_metrics or cls_metrics.get("precision") is None:
                continue
                
            if cls not in aggregators['class_totals']:
                aggregators['class_totals'][cls] = {k: 0.0 for k in aggregators['total_metrics']}
                aggregators['class_counts'][cls] = 0
            
            self._accumulate_metrics(aggregators['class_totals'][cls], cls_metrics)
            aggregators['class_counts'][cls] += 1

    def _compute_and_save_final_results(self, aggregators: Dict) -> None:
        """Compute final averages from aggregates and persist reports."""
        scored_count = aggregators['scored_count']
        scored_count_no_missing = aggregators['scored_count_no_missing']
        
        # Compute overall means.
        avg_metrics = self._compute_average_metrics(
            aggregators['total_metrics'], scored_count
        )
        avg_metrics_no_missing = self._compute_average_metrics(
            aggregators['total_metrics_no_missing'], scored_count_no_missing
        )
        
        # Compute category-level means.
        category_avg = self._compute_category_averages(aggregators)
        
        # Compute class-level means.
        class_avg = self._compute_class_averages(aggregators)
        
        # Build unified summary payload.
        merged_avg = {
            "overall": avg_metrics,
            "overall_no_missing": avg_metrics_no_missing,
            "by_category": category_avg,
            "by_class": class_avg,
            "counts": {
                "total": aggregators['total_count'],
                "scored": scored_count,
                "missing_result": len(self.missing_result_ids),
                "scored_no_missing": scored_count_no_missing,
            },
        }
        
        # Write output files.
        self._write_json(self.output_root / "score_avg.json", merged_avg)
        self._write_category_excel(category_avg, class_avg, avg_metrics)
        
        # Print top-line run summary.
        if avg_metrics:
            print_green(
                f"Overall average score (P/R/F1): "
                f"{avg_metrics.get('precision', 0.0):.4f}/"
                f"{avg_metrics.get('recall', 0.0):.4f}/"
                f"{avg_metrics.get('f1', 0.0):.4f}"
            )
            print_green(
                f"Samples used for averages: "
                f"overall={scored_count}, "
                f"overall_no_missing={scored_count_no_missing}"
            )

    def _compute_average_metrics(
        self, 
        totals: Dict[str, float], 
        count: int
    ) -> Optional[Dict[str, float]]:
        """Return rounded averages for a metric-total map, or None when empty."""
        if not count:
            return None
        return {
            **{k: round(v / count, 4) for k, v in totals.items()},
            "count": count
        }

    def _compute_category_averages(self, aggregators: Dict) -> Dict[str, dict]:
        """Compute per-category averaged metrics in a stable display order."""
        category_avg = {}
        
        # Emit known categories first for deterministic tables.
        for category in self.ordered_categories:
            if category not in aggregators['category_totals']:
                continue
            count = aggregators['category_counts'].get(category, 0)
            if not count:
                continue
            totals = aggregators['category_totals'][category]
            category_avg[category] = {
                **{k: round(v / count, 4) for k, v in totals.items()},
                "count": count,
            }
        
        # Append any extra categories not present in the predefined list.
        for category, totals in aggregators['category_totals'].items():
            if category in category_avg:
                continue
            count = aggregators['category_counts'].get(category, 0)
            if not count:
                continue
            category_avg[category] = {
                **{k: round(v / count, 4) for k, v in totals.items()},
                "count": count,
            }
        
        return category_avg

    def _compute_class_averages(self, aggregators: Dict) -> Dict[str, dict]:
        """Compute per-class averaged metrics."""
        class_avg = {}
        
        for cls in self.ordered_classes:
            total_cls = aggregators['class_item_counts'].get(cls, 0)
            count = aggregators['class_counts'].get(cls, 0)
            
            # Average class metrics over scored records.
            avg_metrics_cls = {}
            if count and cls in aggregators['class_totals']:
                avg_metrics_cls = {
                    k: round(v / count, 4) 
                    for k, v in aggregators['class_totals'][cls].items()
                }
            
            class_avg[cls] = {
                **avg_metrics_cls,
                "count": total_cls,
            }
        
        return class_avg

    def _write_category_excel(
        self,
        category_avg: Dict[str, dict],
        class_avg: Dict[str, dict],
        avg_metrics: Optional[Dict[str, float]],
    ) -> None:
        """Write an Excel report with category/class and overall score views."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "category_scores"

        # Create header row.
        self._create_excel_header(sheet)
        
        # Fill data rows.
        self._fill_excel_data(sheet, category_avg, class_avg, avg_metrics)
        
        # Persist workbook.
        output_path = self.output_root / f"{self.version}_score.xlsx"
        workbook.save(output_path)

    def _create_excel_header(self, sheet) -> None:
        """Create Excel header columns and row labels."""
        # Keep first cell blank for row labels.
        sheet.cell(row=1, column=1, value="")
        
        # Category columns.
        for idx, category in enumerate(self.ordered_categories, start=2):
            sheet.cell(row=1, column=idx, value=category)
        
        # Category overall column.
        category_overall_col = len(self.ordered_categories) + 2
        sheet.cell(row=1, column=category_overall_col, value="Overall")
        
        # Class columns.
        class_start_col = category_overall_col + 1
        for offset, cls in enumerate(self.ordered_classes):
            sheet.cell(row=1, column=class_start_col + offset, value=cls)
        
        # Class overall column.
        class_overall_col = class_start_col + len(self.ordered_classes)
        sheet.cell(row=1, column=class_overall_col, value="Overall")
        
        # Row labels.
        score_row_label = f"{self.version}_score"
        sheet.cell(row=2, column=1, value=score_row_label)

    def _fill_excel_data(
        self, 
        sheet, 
        category_avg: Dict[str, dict],
        class_avg: Dict[str, dict],
        avg_metrics: Optional[Dict[str, float]]
    ) -> None:
        """Populate Excel score rows."""
        def pct_str(value: float) -> str:
            """Format decimal metric as a percentage string with one decimal place."""
            return f"{round(value * 100, 1):.1f}"

        # Fill category columns.
        for idx, category in enumerate(self.ordered_categories, start=2):
            metrics = category_avg.get(category)
            if not metrics:
                continue
            self._fill_metric_cell(sheet, 2, idx, metrics, pct_str)

        # Fill class columns.
        class_start_col = len(self.ordered_categories) + 3
        for offset, cls in enumerate(self.ordered_classes):
            metrics = class_avg.get(cls)
            col = class_start_col + offset
            if not metrics:
                continue
            self._fill_metric_cell(sheet, 2, col, metrics, pct_str)

        # Fill overall columns.
        if avg_metrics:
            category_overall_col = len(self.ordered_categories) + 2
            class_overall_col = class_start_col + len(self.ordered_classes)
            
            self._fill_metric_cell(sheet, 2, category_overall_col, avg_metrics, pct_str)
            self._fill_metric_cell(sheet, 2, class_overall_col, avg_metrics, pct_str)

    def _fill_metric_cell(self, sheet, row: int, col: int, metrics: Dict, pct_str) -> None:
        """Write one `P/R/F1` metric cell in percentage form."""
        precision = metrics.get("precision", 0.0)
        recall = metrics.get("recall", 0.0)
        f1 = metrics.get("f1", 0.0)
        sheet.cell(
            row=row,
            column=col,
            value=f"{pct_str(precision)}/{pct_str(recall)}/{pct_str(f1)}",
        )

    def _load_dataset(self) -> Dict[str, dict]:
        """Load dataset records from `.json` or `.jsonl` into an index-keyed map."""
        with self.dataset_path.open("r", encoding="utf-8") as f:
            if self.dataset_path.suffix == ".jsonl":
                data = [json.loads(line) for line in f if line.strip()]
            else:
                data = json.load(f)
        return {record["index"]: record for record in data}

    def _parse_gold_checklist(self, record: dict) -> Dict[str, dict]:
        """
        Parse the gold checklist from a dataset record.

        Returns:
            {item_id: {"content": str, "pass": bool, "class": str}}
        """
        gold_items: Dict[str, dict] = {}
        for item in record.get("checklist", []):
            gold_id = str(item["id"])
            gold_items[gold_id] = {
                "content": item["content"],
                "pass": bool(item["pass"]),
                "class": item.get("class"),
            }
        return gold_items
    
    def _parse_pred_checklist(self, result_path: Path) -> Dict[str, dict]:
        """
        Parse predicted checklist items from `result_extracted.md`.

        Expected line format:
            - [x] item_id: description
        where `[x]` means pass and `[ ]` means fail.
        """
        pred_items: Dict[str, dict] = {}
        pattern = re.compile(r"^- \[\s*([xX ])\s*\]\s*(?:\*\*)?([A-Za-z0-9_-]+)(?:\*\*)?:\s*(.+)$")
        
        with result_path.open("r", encoding="utf-8") as f:
            for line in f:
                match = pattern.match(line.strip())
                if match:
                    checked = match.group(1)
                    item_id = match.group(2).strip()
                    desc = match.group(3).strip()
                    pred_items[item_id] = {
                        "content": desc, 
                        "pass": checked.lower() == "x"
                    }
        return pred_items

    def _parse_checklist_md(self, checklist_path: Path) -> Dict[str, dict]:
        """
        Parse checklist template items from `checklist.md`.

        Notes:
            `checklist.md` carries item definitions only and no pass/fail signal.
            Every parsed item is therefore treated as pass=True.
        """
        pred_items: Dict[str, dict] = {}
        pattern = re.compile(r"^- \[\s*([xX ])\s*\]\s*(?:\*\*)?([A-Za-z0-9_-]+)(?:\*\*)?:\s*(.+)$")
        
        with checklist_path.open("r", encoding="utf-8") as f:
            for line in f:
                match = pattern.match(line.strip())
                if match:
                    item_id = match.group(2).strip()
                    desc = match.group(3).strip()
                    # checklist.md has no pass/fail labels; use pass=True by design.
                    pred_items[item_id] = {"content": desc, "pass": True}
        return pred_items

    def _build_direct_matches(
        self,
        gold_items: Dict[str, dict],
        pred_items: Dict[str, dict],
    ) -> List[Tuple[str, Optional[str]]]:
        """
        Build direct alignment pairs by checklist order:
        - zip `pred_ids` and `gold_ids` in their original appearance order
        - treat extra predictions as unmatched (`gold_id = None`)
        """
        gold_ids = list(gold_items.keys())
        pred_ids = list(pred_items.keys())

        match_ids: List[Tuple[str, Optional[str]]] = []
        for pred_id, gold_id in zip(pred_ids, gold_ids):
            match_ids.append((pred_id, gold_id))

        # Mark extra predictions as unmatched.
        if len(pred_ids) > len(gold_ids):
            for pred_id in pred_ids[len(gold_ids):]:
                match_ids.append((pred_id, None))

        return match_ids

    def _build_detailed_matches(
        self,
        match_ids: List[Tuple[str, Optional[str]]],
        gold_items: Dict[str, dict],
        pred_items: Dict[str, dict]
    ) -> List[dict]:
        """Build verbose match records including gold/pred text content."""
        # Convert pair list to `gold_id -> [pred_id, ...]`.
        match_map = defaultdict(list)
        for pred_id, gold_id in match_ids:
            match_map[gold_id].append(pred_id)

        detailed_matches = []
        
        # Emit one block per gold item.
        for gold_id, gold_meta in gold_items.items():
            gold_block = {
                "gold": {
                    "id": gold_id,
                    "text": gold_meta.get("content"),
                },
                "pred": None,
            }
            
            # Attach predictions mapped to this gold item.
            for pred_id in match_map.get(gold_id, []):
                pred_meta = pred_items.get(pred_id, {})
                if gold_block["pred"] is None:
                    gold_block["pred"] = []
                gold_block["pred"].append({
                    "id": pred_id,
                    "text": pred_meta.get("content"),
                })
            
            detailed_matches.append(gold_block)
        
        # Add unmatched predictions as a final block with `gold=None`.
        unmatched_preds = match_map.get(None, [])
        if unmatched_preds:
            gold_block = {"gold": None, "pred": []}
            for pred_id in unmatched_preds:
                pred_meta = pred_items.get(pred_id, {})
                gold_block["pred"].append({
                    "id": pred_id,
                    "text": pred_meta.get("content"),
                })
            detailed_matches.append(gold_block)

        return detailed_matches

    def _compute_metrics(
        self,
        match_ids: List[Tuple[str, Optional[str]]],
        gold_items: Dict[str, dict],
        pred_items: Dict[str, dict],
    ) -> Dict[str, float]:
        """
        Compute precision/recall/F1 from matched gold and prediction statuses.

        Evaluation logic:
        - TP: gold indicates a bug and prediction reports failure.
        - FP: gold indicates no bug but prediction reports failure.
        - FN: gold indicates a bug but prediction reports pass.
        - TN: gold indicates no bug and prediction reports pass.
        """
        # Build `gold_id -> [pred_id, ...]` mapping from match pairs.
        gold_to_preds: Dict[str, List[str]] = {}
        for pred_id, gold_id in match_ids or []:
            if gold_id is not None:
                gold_to_preds.setdefault(gold_id, []).append(pred_id)

        tp, fp, fn, tn = 0, 0, 0, 0
        
        for gold_id, gold_meta in gold_items.items():
            pred_ids = gold_to_preds.get(gold_id)
            
            # Case 1: this gold item is not covered by any prediction.
            if pred_ids is None:
                if gold_meta["pass"]:  # No bug in gold; uncovered counts as TN.
                    tn += 1
                else:  # Bug in gold; uncovered means miss (FN).
                    fn += 1
                continue
            
            # Case 2: the gold item is covered; derive predicted pass/fail.
            # Any failing prediction marks the item as "bug found".
            pred_pass = all(pred_items[pred_id]["pass"] for pred_id in pred_ids)
            
            if gold_meta["pass"]:  # Gold says no bug.
                if pred_pass:  # Correct no-bug prediction.
                    tn += 1
                else:  # False positive.
                    fp += 1
            else:  # Gold says bug exists.
                if pred_pass:  # False negative.
                    fn += 1
                else:  # Correctly detected bug.
                    tp += 1
        
        # Derive scalar metrics.
        precision = tp / (tp + fp) if (tp + fp) else 0.0
        recall = tp / (tp + fn) if (tp + fn) else 0.0
        f1 = 2 * precision * recall / (precision + recall) if (precision + recall) else 0.0
        metric = {
            "precision": precision,
            "recall": recall,
            "f1": f1,
        }
        
        return metric, gold_to_preds

    def _write_json(self, path: Path, payload: dict) -> None:
        """Write a JSON file with UTF-8 encoding and stable pretty formatting."""
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    def _process_record(
        self,
        record_id: str,
        record: dict,
        output_dir: Path,
    ) -> Optional[Dict[str, dict]]:
        """
        Score a single dataset record and persist per-record artifacts.

        Returns:
            {
                "overall": overall metrics,
                "by_class": class-level metrics,
                "match_ids": match pairs,
                "gold_items": parsed gold checklist
            }
        }
        """
        result_path = output_dir / "result_extracted.md"
        checklist_path = output_dir / "checklist.md"
        
        # Handle missing output files.
        missing_result = not result_path.exists()
        if missing_result and not checklist_path.exists():
            print(f"Skipping {record_id}: missing result_extracted.md and checklist.md")
            self.missing_result_ids.append(record_id)
            return self._create_missing_result_bundle(record, output_dir)

        if missing_result:
            # No fallback path: mark as missing and stop early.
            self.missing_result_ids.append(record_id)
            return self._create_missing_result_bundle(record, output_dir)

        if missing_result:
            print(f"Missing result_extracted.md for {record_id}; using checklist.md for scoring")
            self.missing_result_ids.append(record_id)

        # Parse gold checklist and predictions.
        gold_items = self._parse_gold_checklist(record)

        pred_items = self._parse_pred_checklist(result_path)
        match_source = "result"

        # Track tasks where all predicted items failed.
        if pred_items and all(not item.get("pass", False) for item in pred_items.values()):
            self.all_fail_pred_ids.append(record_id)

        # Skip scoring when checklist lengths do not match.
        if len(pred_items) != len(gold_items):
            self.mismatch_item_ids.append(record_id)
            self._create_mismatch_result_bundle(record, output_dir, len(pred_items), len(gold_items))
            return None

        # Build direct matches by checklist order.
        match_ids = self._build_direct_matches(gold_items, pred_items)

        # Persist detailed match artifacts for debugging/inspection.
        detailed_matches = self._build_detailed_matches(match_ids, gold_items, pred_items)
        self._write_json(output_dir / "score_match_ids.json", {
            "matches": match_ids,
            "detailed_matches": detailed_matches,
            "source": match_source,
        })

        # Compute overall metrics.
        metrics, _ = self._compute_metrics(match_ids, gold_items, pred_items)
        if missing_result:
            metrics["missing_result"] = True
        
        # Round floats for stable JSON artifacts.
        rounded_metrics = {
            k: (round(v, 4) if isinstance(v, float) else v) 
            for k, v in metrics.items()
        }

        # Compute class-level metrics (only bug-containing classes are scored).
        class_metrics = self._compute_class_metrics(
            match_ids, gold_items, pred_items
        )

        # Persist per-record score output.
        score_payload = {"overall": rounded_metrics}
        for cls in self.ordered_classes:
            score_payload[cls] = class_metrics.get(cls)
        self._write_json(output_dir / "score.json", score_payload)
        
        print_orange(
            f"Scored {record_id}: "
            f"precision={metrics['precision']:.3f}, "
            f"recall={metrics['recall']:.3f}, "
            f"f1={metrics['f1']:.3f}"
        )
        
        return {
            "overall": rounded_metrics,
            "by_class": class_metrics,
            "match_ids": match_ids,
            "gold_items": gold_items,
        }

    def _create_missing_result_bundle(self, record: dict, output_dir: Path) -> dict:
        """Build and write a zeroed score bundle for missing result files."""
        missing_metrics = {
            "precision": 0.0,
            "recall": 0.0,
            "f1": 0.0,
            "missing_result": True,
        }
        self._write_json(output_dir / "score.json", missing_metrics)
        
        missing_by_class = {
            cls: {"precision": 0.0, "recall": 0.0, "f1": 0.0}
            for cls in self.ordered_classes
        }
        
        return {
            "overall": missing_metrics,
            "by_class": missing_by_class,
            "match_ids": None,
            "gold_items": self._parse_gold_checklist(record),
        }

    def _create_mismatch_result_bundle(self, record: dict, output_dir: Path, pred_len: int, gold_len: int) -> dict:
        """Build and write a default bundle for checklist length mismatches."""
        mismatch_metrics = {
            "precision": 0.0,
            "recall": 0.0,
            "f1": 0.0,
            "mismatch_result": True,
            "pred_count": pred_len,
            "gold_count": gold_len,
        }
        self._write_json(output_dir / "score.json", mismatch_metrics)

        mismatch_by_class = {
            cls: {"precision": None, "recall": None, "f1": None}
            for cls in self.ordered_classes
        }

        return {
            "overall": mismatch_metrics,
            "by_class": mismatch_by_class,
            "match_ids": None,
            "gold_items": self._parse_gold_checklist(record),
        }

    def _compute_class_metrics(
        self,
        match_ids: List[Tuple[str, Optional[str]]],
        gold_items: Dict[str, dict],
        pred_items: Dict[str, dict]
    ) -> Dict[str, dict]:
        """Compute per-class metrics for classes that contain at least one bug."""
        class_metrics: Dict[str, dict] = {}
        
        for cls in self.ordered_classes:
            # Extract gold items of this class.
            gold_subset = {
                gid: item 
                for gid, item in gold_items.items() 
                if item.get("class") == cls
            }
            
            if not gold_subset:
                continue
            
            # Skip classes where gold has no failing item.
            has_bug = any(not info.get("pass", False) for info in gold_subset.values())
            if not has_bug:
                continue
            
            # Reuse core metric computation on class subset.
            cls_metrics, _ = self._compute_metrics(match_ids, gold_subset, pred_items)
            class_metrics[cls] = {
                k: (round(v, 4) if isinstance(v, float) else v) 
                for k, v in cls_metrics.items()
            }
        
        # Backfill unscored classes with None for schema stability.
        for cls in self.ordered_classes:
            if cls not in class_metrics:
                class_metrics[cls] = {
                    "precision": None, 
                    "recall": None, 
                    "f1": None
                }
        
        return class_metrics


def parse_args():
    """Parse command-line arguments for the gold-alignment scorer."""
    parser = argparse.ArgumentParser(
        description="Score agent outputs against gold checklist (direct id alignment)."
    )
    parser.add_argument(
        "--dataset_path", type=str,
        help="Path to the WebProber-Bench dataset JSONL file (each line is a record)."
    )
    parser.add_argument(
        "--output_root", type=str,
        help="Root directory for all generated outputs. Each run will create a versioned subdirectory under this root based on --version."
    )
    parser.add_argument(
        "--version", required=True, type=str,
        help="Version label used to group outputs"
    )

    return parser.parse_args()


def main():
    """Program entry point."""
    args = parse_args()

    # Resolve run-specific paths.
    dataset_path = Path(args.dataset_path)
    output_root = Path(args.output_root) / args.version

    # Execute scoring.
    pipeline = ScoringPipeline_Oracle(
        dataset_path=dataset_path,
        output_root=output_root,
        version=args.version,
    )
    pipeline.run()


if __name__ == "__main__":
    main()
